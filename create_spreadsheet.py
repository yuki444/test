"""Excelスプレッドシート作成ユーティリティ"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import sys


def create_spreadsheet(sheets_data: list[dict], output_path: str = "output.xlsx") -> str:
    """
    シートデータからExcelファイルを生成する。

    sheets_data: [
        {
            "name": "シート名",
            "headers": ["列1", "列2", "列3"],
            "rows": [
                ["値1", "値2", "値3"],
                ...
            ]
        },
        ...
    ]
    """
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for sheet_data in sheets_data:
        ws = wb.create_sheet(title=sheet_data.get("name", "Sheet"))

        headers = sheet_data.get("headers", [])
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, row_data in enumerate(sheet_data.get("rows", []), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].auto_size = True
            max_length = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    wb.save(output_path)
    return output_path


def main():
    if len(sys.argv) < 2:
        sample_data = [
            {
                "name": "売上データ",
                "headers": ["月", "売上", "費用", "利益"],
                "rows": [
                    ["1月", 100000, 60000, 40000],
                    ["2月", 120000, 65000, 55000],
                    ["3月", 95000, 58000, 37000],
                    ["4月", 140000, 70000, 70000]
                ]
            },
            {
                "name": "メンバー一覧",
                "headers": ["名前", "部署", "役職", "入社年"],
                "rows": [
                    ["田中 太郎", "開発部", "エンジニア", 2020],
                    ["鈴木 花子", "デザイン部", "デザイナー", 2019],
                    ["佐藤 一郎", "営業部", "マネージャー", 2018]
                ]
            }
        ]
        output = create_spreadsheet(sample_data, "sample.xlsx")
        print(f"サンプルスプレッドシートを作成しました: {output}")
    else:
        json_path = sys.argv[1]
        output_path = sys.argv[2] if len(sys.argv) > 2 else "output.xlsx"
        with open(json_path, "r", encoding="utf-8") as f:
            sheets_data = json.load(f)
        output = create_spreadsheet(sheets_data, output_path)
        print(f"スプレッドシートを作成しました: {output}")


if __name__ == "__main__":
    main()
